import openpyxl
from django import forms
from django.contrib import admin
from django.contrib.auth import get_user_model
from django.core import signing
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import path, reverse
from django.utils.html import format_html

from .models import Location, LocationEntry, Map, Region

# Must match SHARE_MAP_SALT in views.py
_SHARE_MAP_SALT = "travelregistration.share_map"


class UserSelectForm(forms.Form):
    user = forms.ModelChoiceField(
        queryset=get_user_model().objects.order_by('username'),
        label="User",
    )


class ExcelImportForm(forms.Form):
    name = forms.CharField(label="Map name (native)")
    name_en = forms.CharField(label="Map name (English)")
    slug = forms.SlugField()
    excel_file = forms.FileField(label="Excel file (.xlsx)")

    def clean_slug(self):
        slug = self.cleaned_data["slug"]
        if Map.objects.filter(slug=slug).exists():
            raise forms.ValidationError("A map with this slug already exists.")
        return slug


@admin.register(Map)
class MapAdmin(admin.ModelAdmin):
    change_list_template = "admin/travelregistration/map/change_list.html"
    list_display = ("name_en", "slug", "share_link_column")

    def share_link_column(self, obj):
        url = reverse("admin:travelregistration_map_generate_share", args=[obj.pk])
        return format_html('<a class="button" href="{}">Generate share link</a>', url)
    share_link_column.short_description = "Share link"

    def get_urls(self):
        custom = [
            path(
                "import-excel/",
                self.admin_site.admin_view(self.import_excel_view),
                name="travelregistration_map_import_excel",
            ),
            path(
                "<int:map_pk>/generate-share/",
                self.admin_site.admin_view(self.generate_share_view),
                name="travelregistration_map_generate_share",
            ),
        ]
        return custom + super().get_urls()

    def generate_share_view(self, request, map_pk):
        map_obj = get_object_or_404(Map, pk=map_pk)
        share_url = None
        form = UserSelectForm(request.POST or None)
        if request.method == "POST" and form.is_valid():
            user = form.cleaned_data["user"]
            token = signing.dumps({"user": user.pk, "map": map_obj.pk}, salt=_SHARE_MAP_SALT)
            share_url = request.build_absolute_uri(reverse("shared_map", args=[token]))
        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "map": map_obj,
            "form": form,
            "share_url": share_url,
            "title": f"Generate share link — {map_obj.name_en}",
        }
        return render(request, "admin/travelregistration/map/generate_share.html", context)

    def import_excel_view(self, request):
        if request.method == "POST":
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                count = self._import_from_excel(form)
                name_en = form.cleaned_data["name_en"]
                self.message_user(request, f"Imported '{name_en}' with {count} locations.")
                return redirect("admin:travelregistration_map_changelist")
        else:
            form = ExcelImportForm()

        context = {
            **self.admin_site.each_context(request),
            "form": form,
            "opts": self.model._meta,
            "title": "Import map from Excel",
        }
        return render(request, "admin/travelregistration/map/import_excel.html", context)

    def _import_from_excel(self, form):
        wb = openpyxl.load_workbook(form.cleaned_data["excel_file"])
        ws = wb.active

        merge_spans = {}
        merged_non_anchors = set()
        for rng in ws.merged_cells.ranges:
            merge_spans[(rng.min_row, rng.min_col)] = (
                rng.max_col - rng.min_col + 1,
                rng.max_row - rng.min_row + 1,
            )
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    if (r, c) != (rng.min_row, rng.min_col):
                        merged_non_anchors.add((r, c))

        locations = []
        for row in ws.iter_rows():
            for cell in row:
                if (cell.row, cell.column) in merged_non_anchors:
                    continue
                if cell.value is None:
                    continue
                w, h = merge_spans.get((cell.row, cell.column), (1, 1))
                locations.append(Location(
                    name=str(cell.value),
                    name_en=str(cell.value),
                    display_x=cell.column,
                    display_y=cell.row,
                    display_width=w,
                    display_height=h,
                ))

        # Build a set of all occupied (row, col) cells to detect exterior corners.
        occupied = set()
        for loc in locations:
            for r in range(loc.display_y, loc.display_y + loc.display_height):
                for c in range(loc.display_x, loc.display_x + loc.display_width):
                    occupied.add((r, c))

        for loc in locations:
            rc = loc.display_x + loc.display_width - 1   # rightmost column
            br = loc.display_y + loc.display_height - 1  # bottom row
            # A corner is rounded when both adjacent sides at that corner are open.
            loc.border_radius_top_left = (
                (loc.display_y, loc.display_x - 1) not in occupied and
                (loc.display_y - 1, loc.display_x) not in occupied
            )
            loc.border_radius_top_right = (
                (loc.display_y, rc + 1) not in occupied and
                (loc.display_y - 1, rc) not in occupied
            )
            loc.border_radius_bottom_left = (
                (br, loc.display_x - 1) not in occupied and
                (br + 1, loc.display_x) not in occupied
            )
            loc.border_radius_bottom_right = (
                (br, rc + 1) not in occupied and
                (br + 1, rc) not in occupied
            )

        with transaction.atomic():
            map_obj = Map.objects.create(
                name=form.cleaned_data["name"],
                name_en=form.cleaned_data["name_en"],
                slug=form.cleaned_data["slug"],
            )
            region = Region.objects.create(
                map=map_obj,
                name=form.cleaned_data["name_en"],
                color="#888888",
            )
            for loc in locations:
                loc.region = region
            Location.objects.bulk_create(locations)

        return len(locations)


admin.site.register(Region)
admin.site.register(Location)
admin.site.register(LocationEntry)
